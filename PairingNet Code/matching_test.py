import numpy as np
import pickle
from hausdorff import hausdorff_distance
from scipy.spatial.distance import directed_hausdorff
import cv2
import seaborn as sns
import matplotlib.pyplot as plt
import math
import pandas as pd
from openpyxl import load_workbook
import openpyxl

def calculate_hausdorff_distance(set1, set2):
    # 将输入转换为numpy数组
    u = np.array(set1)
    v = np.array(set2)
    
    # 计算Hausdorff距离
    hausdorff_distance = max(directed_hausdorff(u, v)[0], directed_hausdorff(v, u)[0])
    
    return hausdorff_distance, directed_hausdorff(u, v)[0], directed_hausdorff(v, u)[0]

def affine_transform(pt, t):
    """pt: [n, 2]"""
    pt = np.matmul(np.hstack((pt, np.ones((len(pt), 1)))), t.T)
    return pt

def e_rmse(intersection_s_trans, intersection_t):
    """
    calculate the e_rmse score the matching result between two fragments.
    for calculating registration recall.
    """
    ermse = np.sqrt(sum(np.linalg.norm(intersection_s_trans-intersection_t, axis=-1)) / len(intersection_t))
    return ermse

def calculate_area_opencv(points):
    # 将点转换为numpy数组
    contour = np.array(points,dtype=np.int32)
    
    # 计算轮廓的面积
    area = cv2.contourArea(contour)
    
    return area

def calculate_average(lst):
    # 计算列表中所有数的总和
    total = sum(lst)
    # 计算列表中的元素个数
    count = len(lst)
    # 计算平均值
    average = total / count
    return average

def write_excel_xlsx(value, k, path, sheet_name):
    
    try:
        data = openpyxl.load_workbook(path)
        sheetnames = data.get_sheet_names()
        if sheet_name not in sheetnames:
            table = data.create_sheet(sheet_name)

        # sheetnames = data.get_sheet_names()
        # table = data.get_sheet_by_name(sheetnames[0])
        
        table = data.get_sheet_by_name(sheet_name)
        # table = data.active
        print(table.title)  # 输出表名
        nrows = table.max_row  # 获得行数
        ncolumns = table.max_column  # 获得列数
        for i in range(1, len(value)+1):
            for j in range(1, len(value[i-1])+1):
                element = (round(value[i-1][j-1], k) if isinstance(value[i-1][j-1], float) else value[i-1][j-1])
                if isinstance(value[i-1][j-1], float) and value[i-1][j-1]<0.01:
                    element = round((value[i-1][j-1]*10000),k)
                table.cell(nrows+i, j).value = element
        data.save(path)
        print("Data appended to xlsx format table successfully !")
        data.close()

    except FileNotFoundError:
        workbook = openpyxl.Workbook()  
        sheet = workbook.active 
        sheet.title = sheet_name  

        index = len(value)
        for i in range(0, index):
            for j in range(0, len(value[i])):
                element = (round(value[i][j], k) if isinstance(value[i][j], float) else value[i][j])
                sheet.cell(row=i + 1, column=j + 1, value=str(element))  
        workbook.save(path) 
        print("Writing data to xlsx format table successfully!")

def matching_result(test_data, matching_test):
    n = len(matching_test["pred_transformation"])
    haus_list = []
    haus_list_s2t = []
    haus_list_t2s = []
    valid_nums4 = 0  
    valid_nums2 = 0
    valid_nums6 = 0

    radians_error_list = []
    translation_error_list = []
    normlized_translation_error_list = []

    for i in range(0,n):

        GT_transformation = matching_test["GT_transformation"][i]
        pred_transformation = matching_test["pred_transformation"][i]

        idx_s, idx_t = test_data['GT_pairs'][i]
        s_pcd_origin, t_pcd_origin = test_data['full_pcd_all'][idx_s], test_data['full_pcd_all'][idx_t]
        ind_s_origin, ind_t_origin = test_data['source_ind'][i], test_data['target_ind'][i]

        intersection_s = s_pcd_origin[ind_s_origin].reshape(-1, 2)
        intersection_s_trans = affine_transform(intersection_s, pred_transformation)
        intersection_t = t_pcd_origin[ind_t_origin].reshape(-1, 2)

        

        # hausdorff distance指标
        haus_dist = calculate_hausdorff_distance(intersection_s_trans, intersection_t)
        # haus_list.append(np.log2(haus_dist[0]))
        haus_list.append((haus_dist[0]))
        haus_list_s2t.append(np.log2(haus_dist[1]))
        haus_list_t2s.append(np.log2(haus_dist[2]))

        #ERMSE指标，是阈值，yifan测试得到。
        ermes = e_rmse(intersection_s_trans, intersection_t)  
        if ermes < 2:
            valid_nums2 += 1
        elif ermes < 4:
            valid_nums4 += 1
        elif ermes < 6:
            valid_nums6 += 1
        
        radians_pred = math.atan2(pred_transformation[0][1],pred_transformation[0][0])  # atan2()函数得到弧度
        degrees_pred = math.degrees(radians_pred)  # 将弧度转换为角度
        radians_gt = math.atan2(GT_transformation[0][1],GT_transformation[0][0])  # atan2()函数得到弧度
        degrees_gt = math.degrees(radians_gt)  # 将弧度转换为角度

        radians_error = abs(radians_pred-radians_gt)
        radians_error_list.append(radians_error)

        translation_error = math.sqrt(pow(GT_transformation[0][2] - pred_transformation[0][2], 2) + pow(GT_transformation[1][2] - pred_transformation[1][2], 2))
        translation_error_list.append(translation_error)

        source_area, target_area = calculate_area_opencv(s_pcd_origin), calculate_area_opencv(t_pcd_origin)
        normlized_translation_error_list.append(translation_error/(source_area+target_area))


    registration_recall = valid_nums2 / n + valid_nums4 / n
    Hausdorff_distance = calculate_average(haus_list)
    Radians_error = calculate_average(radians_error_list)
    normalized_Translation_error = calculate_average(normlized_translation_error_list)

    return registration_recall, Radians_error, normalized_Translation_error, Hausdorff_distance

def matching_result_different_index(matching_test,test_data, dataset_select, High_1_4_index, Mid_index, Low_1_4_index):
    haus_list = []
    haus_list_s2t = []
    haus_list_t2s = []

    valid_nums_dict = {
        "High":[0,0,0],
        "Mid":[0,0,0],
        "Low":[0,0,0],
    }

    radians_error_list = []
    translation_error_list = []
    pair_count = len(matching_test["GT_transformation"])
    for i in range(0, pair_count):

        GT_transformation = matching_test["GT_transformation"][i]
        pred_transformation = matching_test["pred_transformation"][i]

        idx_s, idx_t = test_data['GT_pairs'][i]
        s_pcd_origin, t_pcd_origin = test_data['full_pcd_all'][idx_s], test_data['full_pcd_all'][idx_t]
        if dataset_select in [50,390]:
            ind_s_origin, ind_t_origin = test_data['source_ind'][i], test_data['target_ind'][i]
        else:
            ind_s_origin, ind_t_origin = test_data['inter_source_ind'][i], test_data['inter_target_ind'][i]
        intersection_s = s_pcd_origin[ind_s_origin].reshape(-1, 2)
        intersection_s_trans = affine_transform(intersection_s, pred_transformation)
        intersection_t = t_pcd_origin[ind_t_origin].reshape(-1, 2)

        # hausdorff distance指标
        haus_dist = calculate_hausdorff_distance(intersection_s_trans, intersection_t)
        haus_list.append(np.log2(haus_dist[0]))
        haus_list_s2t.append(np.log2(haus_dist[1]))
        haus_list_t2s.append(np.log2(haus_dist[2]))

        #ERMSE指标，是阈值，yifan测试得到。
        ermes = e_rmse(intersection_s_trans, intersection_t)  
        
        
        if i in High_1_4_index:
            if ermes < 2:
                valid_nums_dict["High"][0] += 1
            elif ermes < 4:
                valid_nums_dict["High"][1] += 1
            elif ermes < 6:
                valid_nums_dict["High"][2] += 1
        elif i in Mid_index:
            if ermes < 2:
                valid_nums_dict["Mid"][0] += 1
            elif ermes < 4:
                valid_nums_dict["Mid"][1] += 1
            elif ermes < 6:
                valid_nums_dict["Mid"][2] += 1
        elif i in Low_1_4_index:
            if ermes < 2:
                valid_nums_dict["Low"][0] += 1
            elif ermes < 4:
                valid_nums_dict["Low"][1] += 1
            elif ermes < 6:
                valid_nums_dict["Low"][2] += 1
        
        radians_pred = math.atan2(pred_transformation[0][1],pred_transformation[0][0])  
        radians_gt = math.atan2(GT_transformation[0][1],GT_transformation[0][0])  


        radians_error = abs(radians_pred-radians_gt)
        radians_error_list.append(radians_error)

        translation_error = math.sqrt(pow(GT_transformation[0][2] - pred_transformation[0][2], 2) + pow(GT_transformation[1][2] - pred_transformation[1][2], 2))
        
        source_area, target_area = calculate_area_opencv(s_pcd_origin), calculate_area_opencv(t_pcd_origin)
        translation_error_list.append(translation_error/(source_area+target_area))
        
        

    registration_recall_High = valid_nums_dict["High"][0] / len(High_1_4_index) + valid_nums_dict["High"][1] / len(High_1_4_index)
    registration_recall_Mid = valid_nums_dict["Mid"][0] / len(Mid_index) + valid_nums_dict["Mid"][1] / len(Mid_index), valid_nums_dict["Mid"][2] / len(Mid_index)
    registration_recall_Low = valid_nums_dict["Low"][0] / len(Low_1_4_index) + valid_nums_dict["Low"][1] / len(Low_1_4_index), valid_nums_dict["Low"][2] / len(Low_1_4_index)
    

    radians_list_High = [radians_error_list[i] for i in High_1_4_index]
    radians_list_Mid = [radians_error_list[i] for i in Mid_index]
    radians_list_Low = [radians_error_list[i] for i in Low_1_4_index]

    translation_list_High = [translation_error_list[i] for i in High_1_4_index]
    translation_list_Mid = [translation_error_list[i] for i in Mid_index]
    translation_list_Low = [translation_error_list[i] for i in Low_1_4_index]
    
    

    Radians_error_High = calculate_average(radians_list_High)
    Radians_error_Mid = calculate_average(radians_list_Mid)
    Radians_error_Low = calculate_average(radians_list_Low)

    Translation_error_High = calculate_average(translation_list_High)
    Translation_error_Mid = calculate_average(translation_list_Mid)
    Translation_error_Low = calculate_average(translation_list_Low)

    return [registration_recall_High, Radians_error_High, Translation_error_High], \
    [registration_recall_Mid, Radians_error_Mid, Translation_error_Mid], \
    [registration_recall_Low, Radians_error_Low, Translation_error_Low]

def get_matching_result(test_data, matching_test):
    registration_recall, Radians_error, normalized_Translation_error, HD = matching_result(test_data, matching_test)
    write_data = [registration_recall, Radians_error, normalized_Translation_error, HD]
    write_excel_xlsx([[10*"*"], [exp_name,"matching"], ["RR","RE","NTE","HD"],write_data], 3, ROOT+'matching_result.xlsx', sheetName)

    return

if __name__ == "__main__":


    """data reading"""
    test_path = "./test_set_with_downsample.pkl"
    with open(test_path, 'rb') as gt_file:
        test_data = pickle.load(gt_file)

    decending_index_path = "./decending_sorted_scale_index_390.pkl"
    with open(decending_index_path, 'rb') as gt_file:
        decending_index = pickle.load(gt_file)
    
    # 按照正态分布划分区间
    High_1_4_index = decending_index["H_index"]
    Mid_index = decending_index["M_index"]
    Low_1_4_index = decending_index["L_index"]
    ROOT = "./"
    sheetName = "sheet matching"


    """matching test"""
    exp_name = "matching test"
    mat_path = "./saved_test_exp_data.pkl"
    with open(mat_path, 'rb') as gt_file:
        matching_test = pickle.load(gt_file)
    print(exp_name)
    get_matching_result(test_data, matching_test)

    print("Over")
    