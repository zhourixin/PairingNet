import numpy as np
import pickle
import pandas as pd
from openpyxl import load_workbook
import openpyxl

def ndcg(golden, current, n = -1):
    '''
    golden : 1-0 matrix in ideal sorting
    current : 1-0 matrix under real sorting
    n : top n
    '''
    log2_table = np.log2(np.arange(2, 2000))

    def dcg_at_n(rel, n):
        rel = np.asfarray(rel)[:n]
        dcg = np.sum(np.divide(np.power(2, rel) - 1, log2_table[:rel.shape[0]]))
        return dcg

    ndcgs = []
    for i in range(len(current)):
        k = len(current[i]) if n == -1 else n
        idcg = dcg_at_n(sorted(golden[i], reverse=True), n=k)
        dcg = dcg_at_n(current[i], n=k)
        tmp_ndcg = 0 if idcg == 0 else dcg / idcg
        ndcgs.append(tmp_ndcg)
    SUM_ndcgs = sum(ndcgs)
    return 0. if len(ndcgs) == 0 else sum(ndcgs) / (len(ndcgs))

def feature_searching1(result_matrix, gt_pair):
    """to get the topk searching result from score matrix"""
    idx = np.argsort(-result_matrix, axis=1)
    l = []
    for i in range(len(gt_pair)):
        # if mins < len_all[i] <= maxs:
        l.append(np.argwhere(idx[gt_pair[i][0]] == gt_pair[i][1]))

    result = np.array(l).reshape(-1)

    top1 = (result < 1).sum() / len(l)
    top5 = (result < 5).sum() / len(l)
    top10 = (result < 10).sum() / len(l)
    top20 = (result < 20).sum() / len(l)

    l_t2s = []
    
    for j in range(len(gt_pair)):
        l_t2s.append(np.argwhere(idx[gt_pair[j][1]] == gt_pair[j][0]))
    result_t2s = np.array(l_t2s).reshape(-1)
    top1_t2s = (result_t2s < 1).sum() / len(l_t2s)
    top5_t2s = (result_t2s < 5).sum() / len(l_t2s)
    top10_t2s = (result_t2s < 10).sum() / len(l_t2s)
    top20_t2s = (result_t2s < 20).sum() / len(l_t2s)

    return  top5, top10, top20

def feature_searching2(result_matrix, gt_pair):
    """to get the topk searching result from score matrix"""
    result_matrix = result_matrix + result_matrix.T
    idx = np.argsort(-result_matrix, axis=1)
    l = []
    for i in range(len(gt_pair)):
        # if mins < len_all[i] <= maxs:
        l.append(np.argwhere(idx[gt_pair[i][0]] == gt_pair[i][1]))

    result = np.array(l).reshape(-1)

    top1 = (result < 1).sum() / len(l)
    top5 = (result < 5).sum() / len(l)
    top10 = (result < 10).sum() / len(l)
    top20 = (result < 20).sum() / len(l)

    return  top5, top10, top20
def fearching_searching1_different_index(result_matrix, gt_pair, featire_index):
    """to get the topk searching result from score matrix"""
    idx = np.argsort(-result_matrix, axis=1)
    l = []
    for i in range(len(gt_pair)):
        # if mins < len_all[i] <= maxs:
        if i in featire_index:
            l.append(np.argwhere(idx[gt_pair[i][0]] == gt_pair[i][1]))
    l_t2s = []
    for j in range(len(gt_pair)):
        # if mins < len_all[i] <= maxs:
        if j in featire_index:
            l_t2s.append(np.argwhere(idx[gt_pair[j][1]] == gt_pair[j][0]))


    result = np.array(l).reshape(-1)

    top1 = (result < 1).sum() / len(l)
    top5 = (result < 5).sum() / len(l)
    top10 = (result < 10).sum() / len(l)
    top20 = (result < 20).sum() / len(l)

    result_t2s = np.array(l_t2s).reshape(-1)

    top1_t2s = (result_t2s < 1).sum() / len(l_t2s)
    top5_t2s = (result_t2s < 5).sum() / len(l_t2s)
    top10_t2s = (result_t2s < 10).sum() / len(l_t2s)
    top20_t2s = (result_t2s < 20).sum() / len(l_t2s)

    return  top5, top10, top20

def fearching_searching2_different_index(result_matrix, gt_pair, featire_index):
    """to get the topk searching result from score matrix"""
    result_matrix = result_matrix + result_matrix.T
    idx = np.argsort(-result_matrix, axis=1)
    l = []
    for i in range(len(gt_pair)):
        # if mins < len_all[i] <= maxs:
        if i in featire_index:
            l.append(np.argwhere(idx[gt_pair[i][0]] == gt_pair[i][1]))

    result = np.array(l).reshape(-1)

    top1 = (result < 1).sum() / len(l)
    top5 = (result < 5).sum() / len(l)
    top10 = (result < 10).sum() / len(l)
    top20 = (result < 20).sum() / len(l)

    return  top5, top10, top20


def remove_zero_rows(array):
    # 使用numpy的all和axis参数来检查每一行是否全为0
    mask = np.all(array == 0, axis=1)
    # 使用mask来从array中删除全为0的行
    new_array = array[~mask]
    return new_array


def write_excel_xlsx(value, k, path, sheet_name):
    
    try:
        data = openpyxl.load_workbook(path)
        sheetnames = data.get_sheet_names()
        if sheet_name not in sheetnames:
            table = data.create_sheet(sheet_name)
        
        table = data.get_sheet_by_name(sheet_name)

        print(table.title)  
        nrows = table.max_row  
        ncolumns = table.max_column  
        for i in range(1, len(value)+1):
            for j in range(1, len(value[i-1])+1):
                element = (round(value[i-1][j-1], k) if isinstance(value[i-1][j-1], float) else value[i-1][j-1])
                table.cell(nrows+i, j).value = element
        data.save(path)
        print("Data appended to xlsx format table successfully!")

    except FileNotFoundError:
        workbook = openpyxl.Workbook()  
        sheet = workbook.active  
        sheet.title = sheet_name  

        index = len(value)
        for i in range(0, index):
            for j in range(0, len(value[i])):
                element = (round(value[i][j], k) if isinstance(value[i][j], float) else value[i][j])
                sheet.cell(row=i + 1, column=j + 1, value=str(element))  
        workbook.save(path) 
        print("Writing data to xlsx format table successfully!")


def ndcg_test(sim_matrix, GT_pairs):
    idx = np.argsort(-sim_matrix, axis=1)
    length = sim_matrix.shape[0]
    gt = np.zeros((length, length), dtype=np.uint8) 
    pred = np.zeros((length, length), dtype=np.uint8) 
    l = []
    for i in range(len(GT_pairs)):
        location = np.argwhere(idx[GT_pairs[i][0]] == GT_pairs[i][1])[0][0]
        gt[GT_pairs[i][0],GT_pairs[i][1]] = 1
        pred[GT_pairs[i][0],location] = 1
    
    new_gt = remove_zero_rows(gt)
    new_pred = remove_zero_rows(pred)
    NDCG_result = (ndcg(new_gt, new_pred, 5),ndcg(new_gt, new_pred, 10),ndcg(new_gt, new_pred, 20))

    return NDCG_result

def ndcg_test_different_index(sim_matrix, GT_pairs, featire_index):
    idx = np.argsort(-sim_matrix, axis=1)
    length = sim_matrix.shape[0]
    gt = np.zeros((length, length), dtype=np.uint8) 
    pred = np.zeros((length, length), dtype=np.uint8) 
    l = []
    for i in range(len(GT_pairs)):
        if i in featire_index:
            location = np.argwhere(idx[GT_pairs[i][0]] == GT_pairs[i][1])[0][0]
            gt[GT_pairs[i][0],GT_pairs[i][1]] = 1
            pred[GT_pairs[i][0],location] = 1
    
    new_gt = remove_zero_rows(gt)
    new_pred = remove_zero_rows(pred)
    NDCG_result = (ndcg(new_gt, new_pred, 5),ndcg(new_gt, new_pred, 10),ndcg(new_gt, new_pred, 20))

    return NDCG_result

def write_to_excel(exp_name, sheetName, search_data, High_1_3_index, Mid_index, Low_1_3_index, ROOT, method_name="Ours"):
    sim_matrix = search_data["matrix"]
    GT_pairs = search_data["GT_pairs"]

    if method_name=="Ours":
        feature_searching = feature_searching1
        fearching_searching_different_index = fearching_searching1_different_index
    elif method_name in ["jigsawnet","rule_based"]:
        sim_matrix = sim_matrix + sim_matrix.T
        feature_searching = feature_searching1
        fearching_searching_different_index = fearching_searching1_different_index


    ##################
    # recall@k 指标
    ##################
    recall_K = feature_searching(sim_matrix, GT_pairs)
    write_excel_xlsx([[10*"*"], [exp_name,"searching"], ["recall_5","recall_10","recall_20"],recall_K], 3, ROOT+'searching_result.xlsx', sheetName)
    
    ##################
    # NDCG 指标
    ##################
    NDCG_result = ndcg_test(sim_matrix, GT_pairs)
    write_excel_xlsx([["NDCG_5","NDCG_10","NDCG_20"],NDCG_result], 3, ROOT+'searching_result.xlsx', sheetName)

    #（2）不同难度划分的searching测试
    recall_K_high = fearching_searching_different_index(sim_matrix, GT_pairs, High_1_3_index)
    write_excel_xlsx([["High searching"], ["recall_5","recall_10","recall_20"],recall_K_high], 3, ROOT+'searching_result.xlsx', sheetName)
    NDCG_result_High = ndcg_test_different_index(sim_matrix, GT_pairs,High_1_3_index)
    write_excel_xlsx([["NDCG_5","NDCG_10","NDCG_20"],NDCG_result_High], 3, ROOT+'searching_result.xlsx', sheetName)

    recall_K_mid = fearching_searching_different_index(sim_matrix, GT_pairs, Mid_index)
    write_excel_xlsx([["Mid searching"], ["recall_5","recall_10","recall_20"],recall_K_mid], 3, ROOT+'searching_result.xlsx', sheetName)
    NDCG_result_mid = ndcg_test_different_index(sim_matrix, GT_pairs, Mid_index)
    write_excel_xlsx([["NDCG_5","NDCG_10","NDCG_20"],NDCG_result_mid], 3, ROOT+'searching_result.xlsx', sheetName)

    recall_K_low = fearching_searching_different_index(sim_matrix, GT_pairs, Low_1_3_index)
    write_excel_xlsx([["Low searching"], ["recall_5","recall_10","recall_20"],recall_K_low], 3, ROOT+'searching_result.xlsx', sheetName)
    NDCG_result_low = ndcg_test_different_index(sim_matrix, GT_pairs,Low_1_3_index)
    write_excel_xlsx([["NDCG_5","NDCG_10","NDCG_20"],NDCG_result_low], 3, ROOT+'searching_result.xlsx', sheetName)
    
    print(1)

    return


if __name__ == "__main__":


    decending_index_path = "/home/zhourixin/EXP_folder/saved_index/decending_sorted_scale_index_390.pkl"
    with open(decending_index_path, 'rb') as gt_file:
        decending_index = pickle.load(gt_file)
    High_1_3_index = decending_index["H_index"]
    Mid_index = decending_index["M_index"]
    Low_1_3_index = decending_index["L_index"]
    

    
    ROOT = "./"
    
    mat_path = "./sim_matrix_390.pkl"
    with open(mat_path, 'rb') as gt_file:
        searching_test = pickle.load(gt_file)
    print("#"*20)
    print("test_test_test")
    sheetName = "sheet searching"
    exp_name = "test_test_test"
    write_to_excel(exp_name, sheetName, searching_test, High_1_3_index, Mid_index, Low_1_3_index, ROOT)
